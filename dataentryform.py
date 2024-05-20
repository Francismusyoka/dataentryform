import tkinter
from tkcalendar import DateEntry
from tkinter import ttk
from tkinter import messagebox
from os.path import isfile
import openpyxl
import datetime

id_entry = None
date_entry = None
departure_entry = None

def enter_data():
    firstname = first_name_entry.get()
    lastname = last_name_entry.get()
    gender = gender_combobox.get()
    phone = phone_number_entry.get()
    idno = national_id_entry.get()
    department = department_combobox.get()
    date = date_entry.get()
    reptime = report_time_entry.get()
    deptime = depart_time_entry.get()
    
    # creating the excel file 
    if isfile('data_entry.xlsx'):
        pass
    else:
        wb = openpyxl.Workbook()
        ws = wb.create_sheet("Sheet1")
        ws.append(['FName', 'LName', 'Gender', 'ID Number', 'Phone No.', 'Department', 'Date', 'Report Time', 'Depart Time'])
        wb.save("data_entry.xlsx")
    
    wb = openpyxl.load_workbook("data_entry.xlsx")
    ws = wb['Sheet1']
    #check dates to separate reporting for different dates
    if ws.max_row > 1:
        last_date = ws.cell(row=ws.max_row, column=7).value  # Get the date from the last row
        # Compare the last date with the current date entered by the user
        if last_date != date:
            ws.append(['', '', '', '', '', '', '', '', ''])  # Add an empty row as separator for new date
            ws.append([firstname, lastname, gender, idno, phone, department, date, reptime, deptime])  # Append new data
        else:
            ws.append([firstname, lastname, gender, idno, phone, department, date, reptime, deptime])  # Append data to the same date
    else:
        ws.append([firstname, lastname, gender, idno, phone, department, date, reptime, deptime])  # Append data if Excel is empty
        
    
    #ws.append([firstname, lastname, gender, idno, phone, department, date, reptime, deptime])
    wb.save("data_entry.xlsx")
    
    
    
    if firstname and lastname and idno and date and reptime:  
      print("Name: ", firstname, lastname, "ID Number: ", idno)
      print("Gender: ", gender, "Department: ", department,"Phone Number", phone)
      print("Date: ", date, "Report Time: ", reptime, "Departure Time: ", deptime)
    else:
        tkinter.messagebox.showwarning(title = "Error", message = "Please enter all the details!")
        
    clear_fields()
    
def clear_fields():
    first_name_entry.delete(0, tkinter.END)
    last_name_entry.delete(0, tkinter.END)
    gender_combobox.delete(0, tkinter.END)
    phone_number_entry.delete(0, tkinter.END)
    national_id_entry.delete(0, tkinter.END)
    depart_time_entry.delete(0, tkinter.END)
    date_entry.delete(0, tkinter.END)
    report_time_entry.delete(0, tkinter.END)
    depart_time_entry.delete(0  , tkinter.END)
    
#allow only numerics to be entered into phone and id fields
def validate_numeric_input(action, value_if_allowed):
    if action == "1": 
        if value_if_allowed.isdigit() or value_if_allowed == "":
            return True
        else:
            return False
    else:
        return True

#set length of id field to 8 digits
def validate_id_number(action, value_if_allowed):
    if action == "1":  # insert
        if value_if_allowed.isdigit() and len(value_if_allowed) <= 8:
            return True
        else:
            return False
    else:
        return True
#phone number == 10 digits
def validate_phone_number(action, value_if_allowed):
    if action == "1":  # insert
        if value_if_allowed.isdigit() and len(value_if_allowed) <= 10:
            return True
        else:
            return False
    else:
        return True
    
    
#update the departure time in new window
def update_departure_time_window():
    
    global id_entry, date_entry, departure_entry
    
    update_window = tkinter.Toplevel(window)
    update_window.title("Update Departure Time")
    


    # Labels and entry fields for ID number, date, and new departure time
    id_label = tkinter.Label(update_window, text="ID Number:")
    id_label.grid(row=0, column=0, padx=5, pady=5)
    id_entry = tkinter.Entry(update_window)
    id_entry.grid(row=0, column=1, padx=5, pady=5)

    date_label = tkinter.Label(update_window, text="Date:")
    date_label.grid(row=1, column=0, padx=5, pady=5)
    current_date = datetime.date.today() #set date to current date
    date_entry = DateEntry(update_window, selectmode='day', year=current_date.year, month=current_date.month, day=current_date.day)
    date_entry.grid(row=1, column=1, padx=5, pady=5)

    departure_label = tkinter.Label(update_window, text="Departure Time:")
    departure_label.grid(row=2, column=0, padx=5, pady=5)
    departure_entry = tkinter.Entry(update_window)
    departure_entry.grid(row=2, column=1, padx=5, pady=5)
    
    # Button to update departure time
    update_button = tkinter.Button(update_window, text="Update Departure Time", command=lambda: update_departure_time(id_entry.get(), date_entry.get(), departure_entry.get(), update_window))
    update_button.grid(row=3, column=0, columnspan=2, padx=5, pady=5)
    
def update_departure_time(id_number, date, new_departure_time, update_window):
    
    #new_departure_time = departure_entry.get()
    # Load workbook and sheet
    wb = openpyxl.load_workbook("data_entry.xlsx")
    ws = wb['Sheet1']

    # Search for the entry in the Excel file based on ID number and date
    found = False
    for row in ws.iter_rows(values_only=True):
       if row[3] == id_number and row[6] == date:
           row_index = row[0]  # Directly use the row from the iteration
           ws.cell(row=row_index, column=9).value = new_departure_time
           found = True
           break

    if found:
        wb.save("data_entry.xlsx")
        messagebox.showinfo("Success", "Departure time updated successfully.")
    else:
        messagebox.showerror("Error", "Entry not found.")

    # Close the update window
    update_window.destroy()




window = tkinter.Tk()
window.title("Data Entry Form")

frame = tkinter.Frame(window)
frame.pack()

#saving user information
user_info_frame = tkinter.LabelFrame(frame, text="User Information")
user_info_frame.grid(row= 0, column=0, padx=20, pady= 10)

first_name_frame = tkinter.Label(user_info_frame, text="First Name")
first_name_frame.grid(row= 0, column=0)
last_name_frame = tkinter.Label(user_info_frame, text="Last Name")
last_name_frame.grid(row=0, column=1)

first_name_entry = tkinter.Entry(user_info_frame)
last_name_entry = tkinter.Entry(user_info_frame)
first_name_entry.grid(row=1, column=0)
last_name_entry.grid(row=1, column = 1)

gender_label = tkinter.Label(user_info_frame, text = "Gender")
gender_combobox = ttk.Combobox(user_info_frame, values=["Male", "Female"])
gender_label.grid(row=0, column=2)
gender_combobox.grid(row=1, column=2)

national_id_frame = tkinter.Label(user_info_frame, text="National ID Number")
national_id_frame.grid(row= 2, column=0)
national_id_entry = tkinter.Entry(user_info_frame)
national_id_entry.grid(row= 3, column=0)
national_id_entry.config(validate="key", validatecommand=(window.register(validate_numeric_input), "%d", "%P"))
national_id_entry.config(validate="key", validatecommand=(window.register(validate_id_number), "%d", "%P"))

department_label = tkinter.Label(user_info_frame, text = "Department")
department_combobox = ttk.Combobox(user_info_frame, values=["", "ICT", "Finance", "Engineering", "Social work", "Communication", "Procurement", "Education"])
department_label.grid(row=2, column=1)
department_combobox.grid(row=3, column=1)

phone_number_frame = tkinter.Label(user_info_frame, text="Phone Number")
phone_number_frame.grid(row= 2, column=2)
phone_number_entry = tkinter.Entry(user_info_frame)
phone_number_entry.grid(row= 3, column=2)
phone_number_entry.config(validate="key", validatecommand=(window.register(validate_numeric_input), "%d", "%P"))
phone_number_entry.config(validate="key", validatecommand=(window.register(validate_phone_number), "%d", "%P"))


date_frame = tkinter.Label(user_info_frame, text="Date")
date_frame.grid(row= 4, column=0)
current_date = datetime.date.today() #set date to current date
date_entry = DateEntry(user_info_frame, selectmode='day', year=current_date.year, month=current_date.month, day=current_date.day)
date_entry.grid(row=5, column=0)

report_time_frame = tkinter.Label(user_info_frame, text="Report Time")
report_time_frame.grid(row= 4, column=1)
report_time_entry = tkinter.Entry(user_info_frame)
report_time_entry.grid(row=5, column=1)

depart_time_frame = tkinter.Label(user_info_frame, text="Depart Time")
depart_time_frame.grid(row= 4, column=2)
depart_time_entry = tkinter.Entry(user_info_frame)
depart_time_entry.grid(row=5, column=2)

for widget in user_info_frame.winfo_children():
    widget.grid_configure(padx=10, pady = 5)

btns = tkinter.LabelFrame(frame)
btns.grid(row= 1, column=0, padx=20, pady= 10)
    
button = tkinter.Button(btns, text="Submit Details", command = enter_data)
button.grid(row=1, column=0, padx=20, pady= 10)

button = tkinter.Button(btns, text="Update departure time", command = update_departure_time_window)
button.grid(row=1, column=1, padx=20, pady= 10)


window.mainloop()