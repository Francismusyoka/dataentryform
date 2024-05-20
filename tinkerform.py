import tkinter as tk
from tkinter import messagebox
from os.path import isfile
import openpyxl

def save_product():
    product_name = name_entry.get()
    product_description = description_entry.get()
    product_price = price_entry.get()
    
    if isfile('my_workbook.xlsx'):
        pass
    else:
        wb = openpyxl.Workbook()
        ws = wb.create_sheet("Sheet1")
        ws.append(['product_name', 'product_description', 'product_price'])
        wb.save("my_workbook.xlsx")
        
    wb = openpyxl.load_workbook("my_workbook.xlsx")
    ws = wb['Sheet1']
    ws.append([product_name, product_description, product_price])
    wb.save("my_workbook.xlsx")
    
    cell = ws["A1"]
    value = cell.value
    print(value)

    confirmation = messagebox.askquestion("Confirmation", "Are you sure you want to save?")

    if confirmation == 'yes':
        print("Product Name:", product_name)
        print("Product Description:", product_description)
        print("Product Price:", product_price)

product_window = tk.Tk()
product_window.title("Product Details")
product_window.geometry('500x500')
product_window.config(background="Red")


name_label = tk.Label(product_window, text="Product Name:")
name_label.pack()

name_entry = tk.Entry(product_window)
name_entry.pack()

description_label = tk.Label(product_window, text="Product Description:")
description_label.pack()

description_entry = tk.Entry(product_window)
description_entry.pack()

price_label = tk.Label(product_window, text="Product Price:")
price_label.pack()

price_entry = tk.Entry(product_window)
price_entry.pack()

save_button = tk.Button(product_window, text="Save", command=save_product)
save_button.pack(pady=10)

product_window.mainloop()

